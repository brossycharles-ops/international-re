import http.server
import json
import os
import re
from datetime import date
from urllib.parse import urlparse
import openpyxl

PORT = 3000

# Always resolve to the main project root, even if running from inside a worktree
_script_dir = os.path.dirname(os.path.abspath(__file__))
if '.claude' + os.sep + 'worktrees' in _script_dir:
    # Go up 3 levels: <worktree_name> -> worktrees -> .claude -> project_root
    PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(_script_dir)))
else:
    PROJECT_ROOT = _script_dir

PUBLIC_DIR = os.path.join(_script_dir, 'public')
EXCEL_FILE = os.path.join(PROJECT_ROOT, 'newsletter_subscribers.xlsx')


def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Subscribers'
        ws.append(['First Name', 'Last Name', 'Email', 'Date Subscribed'])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 22
        wb.save(EXCEL_FILE)
        print('Created newsletter_subscribers.xlsx')


def handle_subscribe(data):
    first_name = (data.get('firstName') or '').strip()
    last_name = (data.get('lastName') or '').strip()
    email = (data.get('email') or '').strip().lower()

    if not first_name or not last_name or not email:
        return 400, {'error': 'All fields are required.'}

    if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        return 400, {'error': 'Please enter a valid email address.'}

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Subscribers']

    # Check for duplicate
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] and row[2].lower() == email:
            return 409, {'error': 'This email is already subscribed.'}

    ws.append([first_name, last_name, email, str(date.today())])
    wb.save(EXCEL_FILE)

    print(f'New subscriber: {first_name} {last_name} <{email}>')
    return 200, {'message': 'Successfully subscribed!'}


class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=PUBLIC_DIR, **kwargs)

    def do_POST(self):
        if urlparse(self.path).path == '/api/subscribe':
            length = int(self.headers.get('Content-Length', 0))
            try:
                body = json.loads(self.rfile.read(length))
                status, result = handle_subscribe(body)
            except Exception as e:
                print(f'Error: {e}')
                status, result = 500, {'error': 'Server error. Please try again.'}

            self.send_response(status)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(result).encode())
        else:
            self.send_response(404)
            self.end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def log_message(self, format, *args):
        # Suppress noisy static file logs, only show API calls
        if '/api/' in args[0]:
            print(f'  {args[0]} -> {args[1]}')


if __name__ == '__main__':
    init_excel()
    print(f'\n  International RE running at http://localhost:{PORT}\n')
    with http.server.HTTPServer(('', PORT), Handler) as httpd:
        httpd.serve_forever()
